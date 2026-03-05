import os
import re
from typing import List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook


# This module reproduces the Muck Rack ➜ Agility transformation from the legacy
# notebook, preserving mapping and URL extraction behavior.

MAPPING = {
    "Headline": "Article",
    "Outlet": "Media Outlet",
    "Published Date": "Published",
    "Contextual Snippet": "Snippet",
    "Impressions": "UVM (Insights by Similarweb)",
    "Country": "Media Outlet Country",
    "AVE(USD)": "Advertising Value Equivalency",
    "URL": "URL",  # handled specially
}


def _find_header_col_idx(ws, header_name: str) -> Optional[int]:
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    for cell in header_row:
        val = (str(cell.value).strip() if cell.value is not None else "")
        if val.lower() == header_name.lower():
            return cell.column
    return None


def _extract_urls_with_debug(fname: str, n_rows_expected: int, debug: bool = False, debug_limit: int = 12) -> Tuple[List[Optional[str]], List[dict]]:
    """
    Extract URLs from the Articles sheet's URL column:
      - true hyperlinks (cell.hyperlink.target)
      - =HYPERLINK("url", ...) formulas
      - plain text containing http(s)://
    """
    # normal mode to keep hyperlink attributes
    wb_formula = load_workbook(fname, data_only=False)
    ws_formula = wb_formula["Articles"]

    wb_values = load_workbook(fname, data_only=True)
    ws_values = wb_values["Articles"]

    url_idx = _find_header_col_idx(ws_formula, "URL")
    debug_rows = []
    if url_idx is None:
        return [None] * n_rows_expected, [{"error": "URL header not found"}]

    urls: List[Optional[str]] = []
    start_row = 2
    end_row = 1 + n_rows_expected
    url_re = re.compile(r"(https?://[^\s\")]+)", re.IGNORECASE)
    hyp_re = re.compile(r'^\s*=\s*HYPERLINK\(\s*"([^"]+)"', re.IGNORECASE)
    hyp_re_semicolon = re.compile(r'^\s*=\s*HYPERLINK\(\s*"([^"]+)"\s*[,;]', re.IGNORECASE)

    for r in range(start_row, end_row + 1):
        cf = ws_formula.cell(row=r, column=url_idx)  # full cell with formula & hyperlink
        cv = ws_values.cell(row=r, column=url_idx)   # evaluated value
        formula_val = cf.value
        display_val = cv.value
        hl_target = cf.hyperlink.target if cf.hyperlink else None

        extracted = None
        reason = None

        if hl_target:
            extracted = hl_target
            reason = "hyperlink.target"

        if extracted is None and isinstance(formula_val, str) and formula_val.lstrip().startswith("="):
            m = hyp_re_semicolon.search(formula_val) or hyp_re.search(formula_val)
            if m:
                extracted = m.group(1)
                reason = "HYPERLINK(formula)"

        if extracted is None:
            if isinstance(display_val, str):
                m2 = url_re.search(display_val)
                if m2:
                    extracted = m2.group(1)
                    reason = "display text regex"
            if extracted is None and isinstance(formula_val, str):
                m3 = url_re.search(formula_val)
                if m3:
                    extracted = m3.group(1)
                    reason = "formula regex"

        urls.append(extracted)

        if debug and len(debug_rows) < debug_limit:
            preview = lambda s: (s[:120] + "…") if isinstance(s, str) and len(s) > 120 else s
            debug_rows.append({
                "row": r,
                "hyperlink_target": hl_target,
                "formula_val_preview": preview(formula_val),
                "display_val_preview": preview(display_val),
                "extracted_url": extracted,
                "method": reason,
            })

    return urls, debug_rows


def transform_file(input_path: str, output_path: Optional[str] = None, show_debug: bool = False) -> str:
    """
    Transform a single Muck Rack Excel file into Agility format.

    - Reads company name from Summary!B1 (second column header)
    - Maps columns to Agility names
    - Extracts robust URLs from hyperlinks/formulas
    - Writes an Excel file with sheet "Articles"
    """
    if not os.path.exists(input_path):
        raise FileNotFoundError(input_path)

    # Company from Summary header
    summary_header = pd.read_excel(input_path, sheet_name="Summary", nrows=0)
    company_name = summary_header.columns[1] if summary_header.shape[1] > 1 else None

    # Load Articles
    df = pd.read_excel(input_path, sheet_name="Articles")

    required_cols = list(MAPPING.values())
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"Missing columns in input: {missing}")

    out = df.copy()

    # Robust URL extraction
    urls, _debug_rows = _extract_urls_with_debug(
        fname=input_path, n_rows_expected=df.shape[0], debug=show_debug, debug_limit=12
    )

    for agility_col, muck_col in MAPPING.items():
        if agility_col == "Published Date":
            out[agility_col] = pd.to_datetime(df[muck_col], errors="coerce").dt.date
        elif agility_col == "URL":
            out[agility_col] = urls
        else:
            out[agility_col] = df[muck_col]

    out["company"] = company_name

    if output_path is None:
        base_name = os.path.splitext(os.path.basename(input_path))[0]
        output_path = os.path.join(os.path.dirname(input_path), f"transformed/Transformed_{base_name}.xlsx")

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    # Save as 'Raw Data' to be compatible with legacy agility merge step
    out.to_excel(output_path, sheet_name="Raw Data", index=False)
    return output_path

