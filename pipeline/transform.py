"""
Transform Muck Rack Excel files to unified format.
Handles both Muck Rack format (Headline, Outlet) and Agility format (Article, Media Outlet).
"""

import re
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

# Column mapping: Muck Rack name -> unified name
MAPPING = {
    "Headline": "Article",
    "Outlet": "Media Outlet",
    "Published Date": "Published",
    "Contextual Snippet": "Snippet",
    "Impressions": "UVM (Insights by Similarweb)",
    "Country": "Media Outlet Country",
    "AVE(USD)": "Advertising Value Equivalency",
    "URL": "URL",
}
# Reverse: Agility name -> unified name (same in this case)
UNIFIED_COLS = ["Article", "Media Outlet", "Published", "Snippet", "URL", "Media Outlet Country", "Advertising Value Equivalency"]
UNIFIED_COLS_OPTIONAL = ["Author", "Sentiment", "UVM (Insights by Similarweb)"]


def _find_header_col_idx(ws, header_name: str) -> Optional[int]:
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        val = (str(cell.value).strip() if cell.value is not None else "")
        if val.lower() == header_name.lower():
            return cell.column
    return None


def _extract_urls(fname: str, n_rows: int) -> list:
    """Extract URLs from hyperlinks, formulas, or plain text."""
    try:
        wb_formula = load_workbook(fname, data_only=False)
        wb_values = load_workbook(fname, data_only=True)
    except Exception:
        return [None] * n_rows

    ws_formula = wb_formula["Articles"]
    ws_values = wb_values["Articles"]

    url_idx = _find_header_col_idx(ws_formula, "URL")
    if url_idx is None:
        return [None] * n_rows

    url_re = re.compile(r"(https?://[^\s\")]+)", re.IGNORECASE)
    hyp_re = re.compile(r'^\s*=\s*HYPERLINK\(\s*"([^"]+)"', re.IGNORECASE)
    urls = []

    for r in range(2, 2 + n_rows):
        cf = ws_formula.cell(row=r, column=url_idx)
        cv = ws_values.cell(row=r, column=url_idx)
        hl_target = cf.hyperlink.target if cf.hyperlink else None
        formula_val = cf.value
        display_val = cv.value

        extracted = None
        if hl_target:
            extracted = hl_target
        elif isinstance(formula_val, str) and formula_val.lstrip().startswith("="):
            m = hyp_re.search(formula_val)
            if m:
                extracted = m.group(1)
        if extracted is None and isinstance(display_val, str):
            m = url_re.search(display_val)
            if m:
                extracted = m.group(1)
        if extracted is None and isinstance(formula_val, str):
            m = url_re.search(formula_val)
            if m:
                extracted = m.group(1)

        urls.append(extracted)

    return urls


# Brand name aliases: map variants to canonical lowercase name
BRAND_ALIASES = {
    "east capital park rae": "park rae",
}


def _get_brand_from_summary(path: Path) -> str:
    """Read brand/company name from Summary sheet (second column header)."""
    try:
        summary = pd.read_excel(path, sheet_name="Summary", nrows=0)
        if summary.shape[1] > 1:
            name = str(summary.columns[1]).strip()
            raw = name.lower() if name and name != "nan" else ""
            return BRAND_ALIASES.get(raw, raw) if raw else ""
    except Exception:
        pass
    # Fallback: use filename stem (e.g. sirin.xlsx -> sirin)
    stem = path.stem.lower()
    return BRAND_ALIASES.get(stem, stem)


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Map Muck Rack or Agility columns to unified names."""
    # Build reverse mapping: Article -> Article, Headline -> Article, etc.
    col_map = {}
    for muck, unified in MAPPING.items():
        if muck in df.columns and unified not in col_map:
            col_map[unified] = muck
        if unified in df.columns and unified not in col_map:
            col_map[unified] = unified

    out = pd.DataFrame()
    for unified, src in col_map.items():
        if src in df.columns:
            out[unified] = df[src]
    return out


def transform_file(path: Path, brand: str) -> Optional[pd.DataFrame]:
    """Transform a single Muck Rack Excel file to unified DataFrame."""
    try:
        df = pd.read_excel(path, sheet_name="Articles")
    except Exception as e:
        print(f"  Error reading {path.name}: {e}")
        return None

    if df.empty:
        return None

    # Normalize columns
    out = _normalize_columns(df)

    # Ensure required columns
    required = ["Article", "URL", "Media Outlet", "Published", "Snippet"]
    missing = [c for c in required if c not in out.columns]
    if missing:
        print(f"  Skipping {path.name}: missing columns {missing}")
        return None

    # Extract URLs robustly
    urls = _extract_urls(str(path), len(df))
    out["URL"] = urls

    # Add brand (lowercase)
    out["brand"] = brand.lower()

    # Copy over optional columns if present
    for col in ["Author", "Sentiment", "UVM (Insights by Similarweb)", "Advertising Value Equivalency", "Media Outlet Country"]:
        if col in df.columns and col not in out.columns:
            out[col] = df[col]

    return out


def transform_and_merge_new_data(new_data_dir: Path) -> Optional[pd.DataFrame]:
    """
    Transform all Muck Rack Excel files in new_data_dir and merge into one DataFrame.
    Returns DataFrame or None if no files found.
    """
    xlsx_files = list(new_data_dir.glob("*.xlsx")) + list(new_data_dir.glob("*.xls"))
    if not xlsx_files:
        return None

    frames = []
    for path in sorted(xlsx_files):
        brand = _get_brand_from_summary(path)
        df = transform_file(path, brand)
        if df is not None and not df.empty:
            frames.append(df)

    if not frames:
        return None

    return pd.concat(frames, ignore_index=True)
